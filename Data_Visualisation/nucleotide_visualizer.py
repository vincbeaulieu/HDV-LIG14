from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap, LogNorm
from matplotlib.widgets import Button
import matplotlib.patches as patches

# Configuration
PATH = "Datasets/HDV-Lib14.xlsx"
NUCLEOTIDE_MAPPING = {"A": 0, "T": 1, "C": 2, "G": 3, "g": 3}
COLOR_PALETTE = ["purple", "blue", "cyan", "green", "lime", "yellow", 
                 "magenta", "red", "orange", "white", "grey", "black"]
INITIAL_NT_COLORS = {"A": "cyan", "T": "blue", "C": "white", "G": "grey"}


def load_excel_data(path):
    """Load nucleotide columns (G-T) and fitness data from Excel file."""
    wb = load_workbook(path)
    ws = wb.active
    
    sequences = []  # Will be list of lists, each containing 14 nucleotides
    fitness1, fitness2 = [], []
    header = None
    position_labels = []  # Headers from columns G-T
    
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=20, values_only=True)):
        if i == 0:
            continue
        if i == 1:
            header = row
            # Get position labels from columns G-T (indices 6-19)
            position_labels = [str(row[j]) if row[j] is not None else str(j-5) 
                             for j in range(6, 20)]
            continue
        if row[0] is None:
            continue
        
        # Read nucleotides from columns G-T (indices 6-19)
        nt_sequence = []
        for j in range(6, 20):  # Columns G through T
            nt = str(row[j]).strip() if row[j] is not None else ""
            nt_sequence.append(nt)
        
        sequences.append(nt_sequence)
        fitness1.append(float(row[1]) if row[1] is not None else np.nan)
        fitness2.append(float(row[3]) if row[3] is not None else np.nan)
    
    return sequences, fitness1, fitness2, header, position_labels


def create_nucleotide_matrix(sequences, mapping):
    """Convert sequences (list of lists) to numeric matrix."""
    num_sequences = len(sequences)
    num_positions = len(sequences[0]) if sequences else 0
    matrix = np.full((num_sequences, num_positions), np.nan, dtype=float)
    
    for i, seq in enumerate(sequences):
        for j, nucleotide in enumerate(seq):
            if nucleotide in mapping:
                matrix[i, j] = mapping[nucleotide]
    
    return matrix


def build_colormap(nt_colors):
    """Build colormap from nucleotide color dictionary."""
    cmap = ListedColormap([nt_colors["A"], nt_colors["T"], 
                           nt_colors["C"], nt_colors["G"]])
    cmap.set_bad("black")
    return cmap


def get_colorbar_labels(nt_colors):
    """Generate colorbar labels with nucleotide letters and current colors."""
    labels = []
    for nt in ["A", "T", "C", "G"]:
        color_name = nt_colors[nt]
        labels.append(f"{nt} ({color_name})")
    return labels


def setup_figure_layout():
    """Create figure and grid layout."""
    fig = plt.figure(figsize=(14, 6))
    gs = fig.add_gridspec(2, 4, height_ratios=[0.85, 0.15], 
                         width_ratios=[14, 1, 0.5, 1],
                         left=0.10, right=0.95, top=0.95, bottom=0.20, 
                         hspace=0.05, wspace=0.3)
    
    ax_nt = fig.add_subplot(gs[0, 0])
    ax_fit1 = fig.add_subplot(gs[0, 1])
    ax_fit2 = fig.add_subplot(gs[0, 3])
    
    return fig, ax_nt, ax_fit1, ax_fit2


def plot_nucleotide_heatmap(ax, mat_nt, nt_colors, position_labels, title):
    """Plot nucleotide heatmap with colorbar."""
    cmap_nt = build_colormap(nt_colors)
    im_nt = ax.imshow(mat_nt, aspect="auto", interpolation="nearest",
                      cmap=cmap_nt, vmin=0, vmax=3)
    
    # Set axes
    ax.set_xticks(range(len(position_labels)))
    ax.set_xticklabels(position_labels, rotation=45, ha='right')
    ax.set_xlabel("Position")
    ax.set_ylabel("Sequence index")
    ax.set_title(title)
    
    # Add colorbar with nucleotide labels
    cbar = plt.colorbar(im_nt, ax=ax, ticks=[0, 1, 2, 3])
    cbar.ax.set_yticklabels(get_colorbar_labels(nt_colors))
    
    return im_nt, cbar


def plot_fitness_heatmap(ax, mat_fit, colormap, label):
    """Plot fitness heatmap with log scale colorbar."""
    # Filter positive values only
    mat_fit_pos = np.where(mat_fit <= 0, np.nan, mat_fit)
    
    # Create heatmap
    im_fit = ax.imshow(mat_fit_pos, aspect="auto", interpolation="nearest",
                       cmap=colormap,
                       norm=LogNorm(vmin=np.nanmin(mat_fit_pos),
                                   vmax=np.nanmax(mat_fit_pos)))
    
    # Set axes
    ax.set_xticks([])
    ax.set_yticks([])
    ax.set_xlabel(label)
    
    # Add colorbar
    cbar = plt.colorbar(im_fit, ax=ax)
    cbar.set_label(label)
    
    return im_fit


def plot_visualization(mat_nt, mat_fit1, mat_fit2, header, position_labels, nt_colors):
    """Create the main visualization with heatmaps."""
    # Setup layout
    fig, ax_nt, ax_fit1, ax_fit2 = setup_figure_layout()
    
    # Plot nucleotide heatmap
    title = header[0] if header else "Nucleotide Colormap"
    im_nt, cbar_nt = plot_nucleotide_heatmap(ax_nt, mat_nt, nt_colors, 
                                             position_labels, title)
    
    # Plot fitness heatmaps
    fit1_label = header[1] if header else "Fitness 1"
    fit2_label = header[3] if header else "Fitness 2"
    
    im_fit1 = plot_fitness_heatmap(ax_fit1, mat_fit1, plt.cm.Reds, fit1_label)
    im_fit2 = plot_fitness_heatmap(ax_fit2, mat_fit2, plt.cm.Blues, fit2_label)
    
    # Store plot objects for updates
    plot_objects = {
        'im_nt': im_nt,
        'im_fit1': im_fit1,
        'im_fit2': im_fit2,
        'ax_nt': ax_nt,
        'ax_fit1': ax_fit1,
        'ax_fit2': ax_fit2,
        'cbar_nt': cbar_nt
    }
    
    return fig, plot_objects


def create_color_selector_buttons(fig, nt_colors, im_nt, cbar_nt):
    """Create interactive color selector buttons for nucleotides."""
    buttons = []
    selection_markers = {}
    
    y0, height, spacing = 0, 0.03, 0.005
    label_width = 0.03  # Width reserved for nucleotide labels
    buttons_start = 0.05  # Where buttons start
    button_width = 0.65 / len(COLOR_PALETTE)
    
    for i, nt in enumerate(["A", "T", "C", "G"]):
        y_pos = y0 + i * (height + spacing)
        
        # Add nucleotide label - center aligned in the label area
        label_x = buttons_start - label_width / 2
        fig.text(label_x, y_pos + height/2, nt, va='center', ha='center', 
                fontsize=10, weight='bold')
        
        for j, color in enumerate(COLOR_PALETTE):
            x_pos = buttons_start + j * button_width
            ax_btn = fig.add_axes([x_pos, y_pos, button_width, height])
            btn = Button(ax_btn, '', color=color, hovercolor=color)
            
            def make_callback(nt_key, col, btn_ax):
                def callback(event):
                    nt_colors[nt_key] = col
                    
                    # Update colormap
                    new_cmap = build_colormap(nt_colors)
                    im_nt.set_cmap(new_cmap)
                    
                    # Update colorbar labels
                    cbar_nt.ax.set_yticklabels(get_colorbar_labels(nt_colors))
                    
                    # Update selection marker
                    if nt_key in selection_markers:
                        selection_markers[nt_key].remove()
                    
                    rect = patches.Rectangle((0, 0), 1, 1, transform=btn_ax.transAxes,
                                            edgecolor='black', facecolor='none', linewidth=2)
                    btn_ax.add_patch(rect)
                    selection_markers[nt_key] = rect
                    
                    plt.draw()
                return callback
            
            btn.on_clicked(make_callback(nt, color, ax_btn))
            buttons.append(btn)
    
    return buttons


def update_nucleotide_plot(plot_objects, mat_nt, position_labels, title):
    """Update nucleotide heatmap with new data."""
    plot_objects['im_nt'].set_data(mat_nt)
    plot_objects['ax_nt'].set_title(title)
    
    # Update x-axis labels
    plot_objects['ax_nt'].set_xticks(range(len(position_labels)))
    plot_objects['ax_nt'].set_xticklabels(position_labels, rotation=45, ha='right')


def update_fitness_plot(im_fit, ax_fit, mat_fit, label):
    """Update fitness heatmap with new data."""
    mat_fit_pos = np.where(mat_fit <= 0, np.nan, mat_fit)
    
    im_fit.set_data(mat_fit_pos)
    im_fit.set_norm(LogNorm(vmin=np.nanmin(mat_fit_pos),
                           vmax=np.nanmax(mat_fit_pos)))
    ax_fit.set_xlabel(label)


def create_control_buttons(fig, nt_colors, plot_objects):
    """Create refresh color and refresh data buttons."""
    # Reset Colors button
    ax_reset_color = fig.add_axes([0.72, 0.08, 0.10, 0.04])
    btn_reset_color = Button(ax_reset_color, 'Reset Colors', 
                             color='lightgray', hovercolor='gray')
    
    # Refresh Data button
    ax_refresh_data = fig.add_axes([0.83, 0.08, 0.10, 0.04])
    btn_refresh_data = Button(ax_refresh_data, 'Refresh Data', 
                              color='lightblue', hovercolor='skyblue')
    
    def reset_colors(event):
        nt_colors.update(INITIAL_NT_COLORS)
        new_cmap = build_colormap(nt_colors)
        plot_objects['im_nt'].set_cmap(new_cmap)
        
        # Update colorbar labels
        plot_objects['cbar_nt'].ax.set_yticklabels(get_colorbar_labels(nt_colors))
        
        plt.draw()
    
    def refresh_data(event):
        # Reload data from Excel
        sequences, fitness1, fitness2, header, position_labels = load_excel_data(PATH)
        mat_nt = create_nucleotide_matrix(sequences, NUCLEOTIDE_MAPPING)
        mat_fit1 = np.array(fitness1).reshape(-1, 1)
        mat_fit2 = np.array(fitness2).reshape(-1, 1)
        
        # Update nucleotide heatmap
        title = header[0] if header else "Nucleotide Colormap"
        update_nucleotide_plot(plot_objects, mat_nt, position_labels, title)
        
        # Update fitness heatmaps
        fit1_label = header[1] if header else "Fitness 1"
        fit2_label = header[3] if header else "Fitness 2"
        
        update_fitness_plot(plot_objects['im_fit1'], plot_objects['ax_fit1'], 
                          mat_fit1, fit1_label)
        update_fitness_plot(plot_objects['im_fit2'], plot_objects['ax_fit2'], 
                          mat_fit2, fit2_label)
        
        # Redraw
        fig.canvas.draw_idle()
        plt.draw()
    
    btn_reset_color.on_clicked(reset_colors)
    btn_refresh_data.on_clicked(refresh_data)
    
    return btn_reset_color, btn_refresh_data


def main():
    """Main execution function."""
    plt.ion()
    
    # Load and process data
    sequences, fitness1, fitness2, header, position_labels = load_excel_data(PATH)
    mat_nt = create_nucleotide_matrix(sequences, NUCLEOTIDE_MAPPING)
    mat_fit1 = np.array(fitness1).reshape(-1, 1)
    mat_fit2 = np.array(fitness2).reshape(-1, 1)
    
    # Create visualization
    nt_colors = INITIAL_NT_COLORS.copy()
    fig, plot_objects = plot_visualization(mat_nt, mat_fit1, mat_fit2, header, 
                                          position_labels, nt_colors)
    
    # Add interactive controls
    color_buttons = create_color_selector_buttons(fig, nt_colors, 
                                                  plot_objects['im_nt'],
                                                  plot_objects['cbar_nt'])
    reset_color_btn, refresh_data_btn = create_control_buttons(fig, nt_colors, plot_objects)
    
    plt.draw()
    plt.show(block=True)


if __name__ == "__main__":
    main()
